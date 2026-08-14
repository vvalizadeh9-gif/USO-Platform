"""Health Check bulk template — download a pre-filled Excel and parse it back.

Lets a subcontractor fill many sites' results offline in Excel instead of
clicking each site in the app. The template shows only each site's requested
technologies as columns; other technologies are omitted per the rule that
only requested techs are ever asked.
"""
from __future__ import annotations

import io

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.health_check import HcAssignment, HcTask
from app.services.tech_parser import parse_technologies

# Result cell columns per technology follow the pattern "<tech> Result",
# "<tech> Reason", "<tech> Comment".
_ALL_TECHS = ("2G", "3G", "4G")
_HEADER_FILL = PatternFill("solid", fgColor="0B8477")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def build_template(assignment: HcAssignment) -> bytes:
    """Return an .xlsx (as bytes) pre-filled with this assignment's sites.

    Columns: Site ID, Site Name, then per requested technology a Result /
    Reason / Comment triple. A site only gets columns for the techs it
    actually requested; unused tech columns are left blank for that row.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "HealthCheck"

    # Determine the union of requested techs across all sites, so the sheet
    # has a stable column layout; per-row we only expect the site's own techs.
    # Subcontractor fills only Result (Normal/NotNormal) and, when NotNormal,
    # a Comment. The problem category is assigned later by a Coordinator/PM,
    # so there is no Reason column here.
    header = ["Site ID", "Province", "Site Type", "Requested Tech"]
    for tech in _ALL_TECHS:
        header += [f"{tech} Result", f"{tech} Comment"]
    ws.append(header)

    for col_idx in range(1, len(header) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT

    for task in assignment.tasks:
        wi = task.work_item
        techs = parse_technologies(wi.requested_technology) if wi else []
        site_code = wi.site.site_code if wi and wi.site else ""
        province = wi.site.province.name if wi and wi.site and wi.site.province else ""
        site_type = wi.site_type if wi else ""
        row = [
            site_code,
            province,
            site_type,
            "".join(techs),
        ]
        for tech in _ALL_TECHS:
            if tech in techs:
                row += ["", ""]  # empty Result/Comment to fill
            else:
                row += ["N/A", ""]  # not requested for this site
        ws.append(row)

    # Reasonable column widths.
    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def _fmt_dt(value) -> str:
    """Human-readable timestamp for spreadsheet cells (blank when missing)."""
    if value is None:
        return ""
    try:
        return value.strftime("%Y-%m-%d %H:%M")
    except AttributeError:
        return str(value)


def _fmt_result(overall) -> str:
    if overall == "Ready":
        return "Ready"
    if overall == "NotReady":
        return "Not Ready"
    return "Pending"


def _style_header_row(ws, row_idx: int, ncols: int) -> None:
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT


def _tech_detail_columns(technologies: list[dict]) -> list:
    """Flatten a task's technology results into Result/Comment cell pairs.

    ``technologies`` is a list of {technology, result, comment}. Produces two
    cells per fixed technology (2G/3G/4G); a tech the site never reported keeps
    blank cells so every export row lines up under the same header.
    """
    by_tech = {t["technology"]: t for t in technologies}
    cells: list = []
    for tech in _ALL_TECHS:
        entry = by_tech.get(tech)
        if entry is None:
            cells += ["", ""]
        else:
            result = entry.get("result") or ""
            display = "Normal" if result == "Normal" else (
                "Not Normal" if result == "NotNormal" else result
            )
            cells += [display, entry.get("comment") or ""]
    return cells


def build_results_export(result_rows: list[dict]) -> bytes:
    """Build an .xlsx of all completed HC results with full per-tech detail.

    ``result_rows`` items carry: site_code, site_type, province,
    assignment_code, contractor_name, overall_result, problem_category,
    reviewed, completed_at, technologies[{technology, result, comment}].
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "HC Results"

    header = [
        "Site ID", "Province", "Site Type", "Assignment", "Subcontractor",
        "Overall Result", "Problem Category", "Validated", "Feedback Date",
    ]
    for tech in _ALL_TECHS:
        header += [f"{tech} Result", f"{tech} Comment"]
    ws.append(header)
    _style_header_row(ws, 1, len(header))

    for r in result_rows:
        row = [
            r.get("site_code") or "",
            r.get("province") or "",
            r.get("site_type") or "",
            r.get("assignment_code") or "",
            r.get("contractor_name") or "",
            _fmt_result(r.get("overall_result")),
            r.get("problem_category") or "",
            "Yes" if r.get("reviewed") else "No",
            _fmt_dt(r.get("completed_at")),
        ]
        row += _tech_detail_columns(r.get("technologies") or [])
        ws.append(row)

    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def build_assignment_feedback_export(
    assignment: HcAssignment,
    *,
    contractor_name: str | None,
    stats: dict,
) -> bytes:
    """Build an .xlsx for a single assignment: summary block + per-site detail.

    ``stats`` is the dict from ``health_check.build_assignment_stats``.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Assignment Feedback"

    aging = stats.get("aging_days")
    aging_txt = (
        "" if aging is None
        else f"{aging} day(s){' (ongoing)' if stats.get('aging_ongoing') else ''}"
    )
    summary = [
        ("Assignment", assignment.code),
        ("Subcontractor", contractor_name or ""),
        ("Assignment Date", _fmt_dt(stats.get("assigned_at"))),
        ("Feedback Received Date", _fmt_dt(stats.get("feedback_received_at"))),
        ("Aging (assignment → feedback)", aging_txt),
        ("Sites Assigned", stats.get("sites_assigned", 0)),
        ("Sites Ready", stats.get("sites_ready", 0)),
        ("Sites Not Ready", stats.get("sites_not_ready", 0)),
        ("Sites Pending", stats.get("sites_pending", 0)),
        ("Status", assignment.status),
    ]
    for label, value in summary:
        ws.append([label, value])
        ws.cell(row=ws.max_row, column=1).font = Font(bold=True)

    ws.append([])  # spacer row before the per-site table

    header = ["Site ID", "Province", "Site Type", "Overall Result",
              "Problem Category", "Validated"]
    for tech in _ALL_TECHS:
        header += [f"{tech} Result", f"{tech} Comment"]
    ws.append(header)
    _style_header_row(ws, ws.max_row, len(header))

    for task in assignment.tasks:
        wi = task.work_item
        site = wi.site if wi else None
        technologies = [
            {"technology": t.technology, "result": t.result, "comment": t.comment}
            for t in task.technologies
        ]
        row = [
            site.site_code if site else (f"#{task.work_item_id}"),
            site.province.name if site and site.province else "",
            wi.site_type if wi else "",
            _fmt_result(task.overall_result),
            task.problem_category or "",
            "Yes" if task.reviewed_at is not None else "No",
        ]
        row += _tech_detail_columns(technologies)
        ws.append(row)

    for col_idx in range(1, len(header) + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 18

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def parse_template(content: bytes, assignment: HcAssignment) -> list[dict]:
    """Parse a filled template into per-task result payloads.

    Returns a list of:
        {work_item_id, technology_results: [{technology, result,
                                             reason_category, comment}]}

    Raises ValueError on validation problems (e.g. NotNormal without reason)
    so the caller can report a clear error and apply nothing.
    """
    df = pd.read_excel(io.BytesIO(content), sheet_name="HealthCheck")

    # Map site code -> task for quick lookup.
    by_code = {}
    for task in assignment.tasks:
        wi = task.work_item
        if wi and wi.site:
            by_code[str(wi.site.site_code)] = task

    results: list[dict] = []
    for _, row in df.iterrows():
        site_id = _clean(row.get("Site ID"))
        if not site_id or site_id not in by_code:
            continue
        task = by_code[site_id]
        wi = task.work_item
        requested = parse_technologies(wi.requested_technology) if wi else []

        tech_results = []
        for tech in requested:
            result = _clean(row.get(f"{tech} Result"))
            if result not in ("Normal", "NotNormal"):
                raise ValueError(
                    f"Site {site_id}, {tech}: result must be 'Normal' or "
                    f"'NotNormal' (got '{result or 'blank'}')."
                )
            comment = _clean(row.get(f"{tech} Comment"))
            if result == "NotNormal" and not comment:
                raise ValueError(
                    f"Site {site_id}, {tech}: 'NotNormal' requires a comment."
                )
            tech_results.append(
                {
                    "technology": tech,
                    "result": result,
                    "comment": comment,
                }
            )

        results.append(
            {"work_item_id": task.work_item_id, "technology_results": tech_results}
        )

    return results


def _clean(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).strip()
    if text in ("", "nan", "NaN", "None", "N/A"):
        return None
    return text
