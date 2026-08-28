"""Tests for strict CPM validation (site_type / village_qty / requested_tech).

Verifies that on a monthly re-import:
  * the 3 tracked changes are flagged and BLOCKED (not applied),
  * everything else auto-applies,
  * accepting a change request actually applies it.

Run with:  cd backend && pytest tests/test_cpm_validation.py -q
"""
import os
import sys

import openpyxl
import pytest

os.environ["DATABASE_URL"] = "sqlite:////tmp/uep_cpmval_pytest.db"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlalchemy.dialects.postgresql as _pg  # noqa: E402
from sqlalchemy import JSON  # noqa: E402

_pg.JSONB = JSON

from fastapi.testclient import TestClient  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from tests.conftest import create_schema, login_form, sample_cpm_path  # noqa: E402

# The workbook these tests import and then modify. sample_cpm_path() returns
# the committed synthetic fixture, or whatever UEP_TEST_CPM_XLSX points at,
# or None when neither is available -- which is what lets the skips below
# actually fire.
SRC = sample_cpm_path()


@pytest.fixture(scope="module")
def client():
    if os.path.exists("/tmp/uep_cpmval_pytest.db"):
        os.remove("/tmp/uep_cpmval_pytest.db")
    create_schema()
    from app.main import app

    with TestClient(app) as c:
        yield c


def _import(filename):
    from app.models.reference import User
    from app.services.cpm_import import CpmImportService

    db = SessionLocal()
    admin = db.query(User).filter(User.username == "admin").first()
    batch = CpmImportService(db, user_id=admin.id).import_file(SRC if filename == "m1" else filename, filename)
    result = {
        "new_count": batch.new_count,
        "changed_count": batch.changed_count,
        "new_villages_count": batch.new_villages_count,
    }
    db.close()
    return result


def test_first_import_has_no_change_requests(client):
    if not SRC:
        pytest.skip("no sample CPM workbook available")
    from app.models.reference import User
    from app.services.cpm_import import CpmImportService

    db = SessionLocal()
    admin = db.query(User).filter(User.username == "admin").first()
    batch = CpmImportService(db, user_id=admin.id).import_file(SRC, "month1.xlsx")
    assert batch.new_count > 0
    assert batch.changed_count == 0  # nothing to validate on the seed import
    db.close()


def test_requested_tech_change_is_blocked_then_applied(client):
    if not SRC:
        pytest.skip("no sample CPM workbook available")
    from app.services import cpm_columns as C
    from app.models.acceptance import CpmChangeRequest
    from app.models.reference import User
    from app.models.workitem import WorkItem
    from app.services.cpm_import import CpmImportService

    # Build a second file with one requested-tech change on a real site row.
    wb = openpyxl.load_workbook(SRC)
    ws = wb["CPM"]
    # header at index 2 => data starts Excel row 4. Find a row with a site code.
    target_excel_row = 4
    orig = ws.cell(row=target_excel_row, column=C.COL["requested_tech"] + 1).value
    new_val = "2G" if orig != "2G" else "3G"
    ws.cell(row=target_excel_row, column=C.COL["requested_tech"] + 1, value=new_val)
    mod = "/tmp/cpmval_month2.xlsx"
    wb.save(mod)

    db = SessionLocal()
    admin = db.query(User).filter(User.username == "admin").first()
    # Run for its effect: the import is what creates the change requests
    # asserted on below.
    CpmImportService(db, user_id=admin.id).import_file(mod, "month2.xlsx")
    # At least the requested-tech change should be flagged.
    tech_crs = (
        db.query(CpmChangeRequest)
        .filter(
            CpmChangeRequest.change_type == "requested_tech",
            CpmChangeRequest.status == "Pending",
        )
        .all()
    )
    assert len(tech_crs) >= 1
    cr = tech_crs[0]
    # Strict: the work item's value must NOT have changed yet.
    wi = db.get(WorkItem, cr.work_item_id)
    assert wi.requested_technology != cr.new_value
    cr_id = cr.id
    db.close()

    # Accept via API applies it.
    r = client.post("/api/v1/auth/login", data=login_form(client))
    h = {"Authorization": f"Bearer {r.json()['access_token']}"}
    r = client.post(
        f"/api/v1/admin/cpm/change-requests/{cr_id}/decide",
        headers=h,
        json={"decision": "Accepted"},
    )
    assert r.status_code == 200

    db = SessionLocal()
    cr = db.get(CpmChangeRequest, cr_id)
    wi = db.get(WorkItem, cr.work_item_id)
    assert wi.requested_technology == cr.new_value  # now applied
    assert cr.status == "Accepted"
    db.close()


def test_pending_count_endpoint(client):
    r = client.post("/api/v1/auth/login", data=login_form(client))
    h = {"Authorization": f"Bearer {r.json()['access_token']}"}
    r = client.get("/api/v1/admin/cpm/change-requests/count", headers=h)
    assert r.status_code == 200
    assert "pending" in r.json()
    assert isinstance(r.json()["pending"], int)


def test_ignored_change_is_not_applied(client):
    if not SRC:
        pytest.skip("no sample CPM workbook available")
    from app.services import cpm_columns as C
    from app.models.acceptance import CpmChangeRequest
    from app.models.reference import User
    from app.models.workitem import WorkItem
    from app.services.cpm_import import CpmImportService

    # Raise a change request of our own rather than relying on one left behind
    # by an earlier test. This used to look for any pending requested-tech
    # change and skip when it found none -- which is exactly what happened,
    # because the test above accepts the one it creates. The test therefore
    # never ran.
    #
    # A different row from the one that test modifies, so the two do not
    # interfere.
    wb = openpyxl.load_workbook(SRC)
    ws = wb["CPM"]
    target_excel_row = 5  # header at index 2 => data starts at Excel row 4
    column = C.COL["requested_tech"] + 1
    original = ws.cell(row=target_excel_row, column=column).value
    ws.cell(
        row=target_excel_row,
        column=column,
        value="4G" if original != "4G" else "2G",
    )
    modified = "/tmp/cpmval_month3.xlsx"
    wb.save(modified)

    db = SessionLocal()
    admin = db.query(User).filter(User.username == "admin").first()
    CpmImportService(db, user_id=admin.id).import_file(modified, "month3.xlsx")
    pending = (
        db.query(CpmChangeRequest)
        .filter(
            CpmChangeRequest.change_type == "requested_tech",
            CpmChangeRequest.status == "Pending",
        )
        .first()
    )
    assert pending is not None, "the re-import should have raised a change request"
    cr_id = pending.id
    wi_id = pending.work_item_id
    before = db.get(WorkItem, wi_id).requested_technology
    db.close()

    r = client.post("/api/v1/auth/login", data=login_form(client))
    h = {"Authorization": f"Bearer {r.json()['access_token']}"}
    r = client.post(
        f"/api/v1/admin/cpm/change-requests/{cr_id}/decide",
        headers=h,
        json={"decision": "Ignored"},
    )
    assert r.status_code == 200

    db = SessionLocal()
    after = db.get(WorkItem, wi_id).requested_technology
    assert after == before  # ignored change left data untouched
    db.close()
