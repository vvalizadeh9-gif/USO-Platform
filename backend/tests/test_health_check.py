"""Tests for the Health Check workflow (Phase A).

Run with:  cd backend && pytest tests/test_health_check.py -q
"""
import os
import sys

import pytest

os.environ["DATABASE_URL"] = "sqlite:////tmp/uep_hc_pytest.db"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlalchemy.dialects.postgresql as _pg  # noqa: E402
from sqlalchemy import JSON  # noqa: E402

_pg.JSONB = JSON

from fastapi.testclient import TestClient  # noqa: E402

from app.core.database import SessionLocal  # noqa: E402
from tests.conftest import create_schema, login_as_role, login_form, sample_cpm_path  # noqa: E402


@pytest.fixture(scope="module")
def client():
    if os.path.exists("/tmp/uep_hc_pytest.db"):
        os.remove("/tmp/uep_hc_pytest.db")
    create_schema()
    from app.main import app

    with TestClient(app) as c:
        yield c


def _login(client):
    r = client.post("/api/v1/auth/login", data=login_form(client))
    return r.json()["access_token"]


def _pm(client):
    """Headers for a PM.

    The health-check workflow is deliberately closed to Admin: assigning is for
    a Coordinator or PM, and submitting results is for a PM or the
    subcontractor. Admin still performs the CPM import below, which is
    admin-only.
    """
    return login_as_role(client, _login(client), "PM")


def _seed(client, token):
    """Import the sample CPM so the basket has on-air sites."""
    sample = sample_cpm_path()
    if not sample:
        pytest.skip("sample CPM file not available")
    with open(sample, "rb") as f:
        client.post(
            "/api/v1/admin/cpm/import",
            headers={"Authorization": f"Bearer {token}"},
            files={"file": ("CPM_2_.xlsx", f, "application/vnd.ms-excel")},
        )


def test_technology_parser():
    from app.services.tech_parser import parse_technologies

    assert parse_technologies("2G3G4G") == ["2G", "3G", "4G"]
    assert parse_technologies("2G4G") == ["2G", "4G"]
    assert parse_technologies("3G") == ["3G"]
    assert parse_technologies(None) == []
    assert parse_technologies("") == []


def test_basket_only_shows_requested_techs(client):
    token = _login(client)
    _seed(client, token)
    r = client.get("/api/v1/hc/basket", headers={"Authorization": f"Bearer {token}"})
    assert r.status_code == 200
    basket = r.json()
    assert len(basket) > 0
    # Every basket item lists only real technologies, never empty/other tokens.
    for item in basket:
        for tech in item["requested_technologies"]:
            assert tech in ("2G", "3G", "4G", "5G")


def test_all_normal_gives_ready(client):
    _login(client)
    h = _pm(client)
    basket = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    site = basket[0]
    r = client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": [site["work_item_id"]]},
    )
    task = r.json()["tasks"][0]

    tech_results = [
        {"technology": t, "result": "Normal"} for t in site["requested_technologies"]
    ]
    r = client.post(
        f"/api/v1/hc/tasks/{task['id']}/result",
        headers=h,
        json={"technology_results": tech_results},
    )
    assert r.status_code == 200
    assert r.json()["overall_result"] == "Ready"


def test_any_not_normal_gives_not_ready_with_category(client):
    _login(client)
    h = _pm(client)
    basket = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    site = basket[0]
    techs = site["requested_technologies"]
    r = client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": [site["work_item_id"]]},
    )
    task = r.json()["tasks"][0]

    results = [
        {
            "technology": techs[0],
            "result": "NotNormal",
            "reason_category": "Temporary Power",
            "comment": "no power on site",
        }
    ]
    for t in techs[1:]:
        results.append({"technology": t, "result": "Normal"})

    r = client.post(
        f"/api/v1/hc/tasks/{task['id']}/result",
        headers=h,
        json={"technology_results": results},
    )
    assert r.status_code == 200
    assert r.json()["overall_result"] == "NotReady"


def test_not_normal_requires_reason_and_comment(client):
    _login(client)
    h = _pm(client)
    basket = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    site = basket[0]
    techs = site["requested_technologies"]
    r = client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": [site["work_item_id"]]},
    )
    task = r.json()["tasks"][0]

    # NotNormal with no reason/comment must be rejected.
    results = [{"technology": techs[0], "result": "NotNormal"}]
    for t in techs[1:]:
        results.append({"technology": t, "result": "Normal"})
    r = client.post(
        f"/api/v1/hc/tasks/{task['id']}/result",
        headers=h,
        json={"technology_results": results},
    )
    assert r.status_code == 400


def test_wrong_technology_rejected(client):
    _login(client)
    h = _pm(client)
    basket = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    # Find a site that does NOT request 5G (none do) and try to submit 5G.
    site = basket[0]
    r = client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": [site["work_item_id"]]},
    )
    task = r.json()["tasks"][0]
    r = client.post(
        f"/api/v1/hc/tasks/{task['id']}/result",
        headers=h,
        json={"technology_results": [{"technology": "5G", "result": "Normal"}]},
    )
    assert r.status_code == 400


def test_assigned_site_leaves_basket(client):
    _login(client)
    h = _pm(client)
    before = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    ids = [before[0]["work_item_id"]]
    client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": ids},
    )
    after = client.get("/api/v1/hc/basket", headers=h).json()
    after_ids = {b["work_item_id"] for b in after}
    assert ids[0] not in after_ids


def test_bulk_template_download_and_upload(client):
    import io

    from openpyxl import load_workbook

    _login(client)
    h = _pm(client)
    basket = client.get("/api/v1/hc/basket", headers=h).json()

    from app.models.reference import Contractor

    db = SessionLocal()
    contractor_id = db.query(Contractor).first().id
    db.close()

    ids = [b["work_item_id"] for b in basket[:2]]
    r = client.post(
        "/api/v1/hc/assignments",
        headers=h,
        json={"contractor_id": contractor_id, "work_item_ids": ids},
    )
    assignment = r.json()

    # Download template
    r = client.get(
        f"/api/v1/hc/assignments/{assignment['id']}/template", headers=h
    )
    assert r.status_code == 200
    wb = load_workbook(io.BytesIO(r.content))
    ws = wb.active
    headers = [c.value for c in ws[1]]

    # Fill every requested tech as Normal
    for row in ws.iter_rows(min_row=2):
        for tech in ("2G", "3G", "4G"):
            col_name = f"{tech} Result"
            if col_name in headers:
                cell = row[headers.index(col_name)]
                if cell.value != "N/A":
                    cell.value = "Normal"
    out = io.BytesIO()
    wb.save(out)

    r = client.post(
        f"/api/v1/hc/assignments/{assignment['id']}/upload",
        headers=h,
        files={
            "file": (
                "filled.xlsx",
                out.getvalue(),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert r.status_code == 200
    assert r.json()["sites_applied"] == len(ids)

    # All should be Ready.
    r = client.get(f"/api/v1/hc/assignments/{assignment['id']}", headers=h)
    for t in r.json()["tasks"]:
        assert t["overall_result"] == "Ready"
