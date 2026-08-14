"""Tests for the هدف-only import gate.

Only villages classified exactly هدف enter the database. Every other value of
the هدف/اقماری column is counted and dropped:

    اقماری
    اقماری جهت بررسی پوشش
    اقماری (Removed Verbally)
    هدف (Verbally)
    هدف (Removed Verbally)

These tests build their own small CPM workbook, so they run without the sample
spreadsheet fixture.

Run with:  cd backend && pytest tests/test_cpm_target_filter.py -q
"""
import os
import sys

import openpyxl
import pytest

os.environ["DATABASE_URL"] = "sqlite:////tmp/uep_target_pytest.db"
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

import sqlalchemy.dialects.postgresql as _pg  # noqa: E402
from sqlalchemy import JSON  # noqa: E402

_pg.JSONB = JSON

from fastapi.testclient import TestClient  # noqa: E402

from app.core.database import Base, SessionLocal, engine  # noqa: E402
from app.services import cpm_columns as C  # noqa: E402

# The five classifications the CPM file uses that must never be imported.
IGNORED_CLASSIFICATIONS = (
    "اقماری جهت بررسی پوشش",
    "اقماری",
    "هدف  (Removed Verbally)",  # double space — as it appears in the file
    "هدف (Verbally)",
    "اقماری (Removed Verbally)",
)


@pytest.fixture(scope="module")
def client():
    if os.path.exists("/tmp/uep_target_pytest.db"):
        os.remove("/tmp/uep_target_pytest.db")
    Base.metadata.drop_all(bind=engine)
    from app.main import app

    with TestClient(app) as c:
        yield c


def _write_cpm(path: str, rows: list[dict]) -> str:
    """Write a minimal CPM workbook: header on row 3, data from row 4."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = C.SHEET_NAME
    last_col = max(C.COL.values()) + 1
    for col in range(1, last_col + 1):
        ws.cell(row=C.HEADER_ROW + 1, column=col, value=f"H{col}")
    for offset, row in enumerate(rows):
        excel_row = C.HEADER_ROW + 2 + offset
        for key, value in row.items():
            ws.cell(row=excel_row, column=C.COL[key] + 1, value=value)
    wb.save(path)
    return path


def _row(site, village, target_flag, **extra):
    row = {
        "site_code": site,
        "site_type": "Macro",
        "village_code": village,
        "village_name": f"Village {village}",
        "province": "تهران",
        "requested_tech": "2G",
        "target_flag": target_flag,
    }
    row.update(extra)
    return row


def _import(path: str, filename: str):
    from app.models.reference import User
    from app.services.cpm_import import CpmImportService

    db = SessionLocal()
    admin = db.query(User).filter(User.username == "admin").first()
    batch = CpmImportService(db, user_id=admin.id).import_file(path, filename)
    summary = {
        "total_rows": batch.total_rows,
        "new_count": batch.new_count,
        "new_villages_count": batch.new_villages_count,
        "skipped": batch.skipped_satellite,
        "changed_village_qty": batch.changed_village_qty,
    }
    db.close()
    return summary


def test_is_pure_target_accepts_only_هدف():
    assert C.is_pure_target("هدف") is True
    assert C.is_pure_target("  هدف  ") is True
    for value in IGNORED_CLASSIFICATIONS:
        assert C.is_pure_target(value) is False, value
    assert C.is_pure_target(None) is False
    assert C.is_pure_target("---") is False


def test_only_هدف_villages_are_imported(client):
    """Seed import: هدف rows land, the five ignored categories are counted."""
    from app.models.workitem import Site, Village, WorkItem

    rows = [
        _row("SITE-A", "V-KEEP-1", "هدف"),
        _row("SITE-A", "V-KEEP-2", " هدف "),  # whitespace variant still counts
    ]
    # One row per ignored category, each on its own site so a leak would also
    # create a site/work item we can assert against.
    rows += [
        _row(f"SITE-SKIP-{i}", f"V-SKIP-{i}", flag)
        for i, flag in enumerate(IGNORED_CLASSIFICATIONS)
    ]
    path = _write_cpm("/tmp/uep_target_seed.xlsx", rows)

    summary = _import(path, "seed.xlsx")
    assert summary["total_rows"] == 7
    assert summary["skipped"] == 5
    assert summary["new_count"] == 1          # only SITE-A / Macro
    assert summary["new_villages_count"] == 2

    db = SessionLocal()
    assert {v.village_code for v in db.query(Village).all()} == {"V-KEEP-1", "V-KEEP-2"}
    assert {v.target_classification for v in db.query(Village).all()} == {"هدف"}
    assert [s.site_code for s in db.query(Site).all()] == ["SITE-A"]
    assert db.query(WorkItem).count() == 1
    db.close()


def test_reimport_downgrade_out_of_هدف_is_a_village_qty_change(client):
    """A village that leaves هدف is dropped from the file, so the re-import
    raises it as a village-quantity change for PM validation rather than
    silently mutating the stored row."""
    from app.models.acceptance import CpmChangeRequest
    from app.models.workitem import Village

    rows = [
        _row("SITE-A", "V-KEEP-1", "هدف"),
        _row("SITE-A", "V-KEEP-2", "هدف (Removed Verbally)"),  # was هدف last month
    ]
    path = _write_cpm("/tmp/uep_target_month2.xlsx", rows)

    summary = _import(path, "month2.xlsx")
    assert summary["skipped"] == 1
    assert summary["changed_village_qty"] == 1

    db = SessionLocal()
    cr = (
        db.query(CpmChangeRequest)
        .filter(CpmChangeRequest.change_type == "village_qty")
        .one()
    )
    assert cr.status == "Pending"
    assert "V-KEEP-2" in cr.detail
    # Strict: nothing is applied until a PM accepts.
    assert db.query(Village).count() == 2
    db.close()


def test_reimport_promotion_into_هدف_refreshes_classification(client):
    """The mirror case: a village stored with a verbal sub-flag by an older
    import is promoted to plain هدف when the file says so."""
    from app.models.workitem import Village

    db = SessionLocal()
    legacy = db.query(Village).filter(Village.village_code == "V-KEEP-2").one()
    legacy.target_classification = "هدف (Verbally)"  # as a laxer import left it
    db.commit()
    db.close()

    rows = [
        _row("SITE-A", "V-KEEP-1", "هدف"),
        _row("SITE-A", "V-KEEP-2", "هدف"),
    ]
    path = _write_cpm("/tmp/uep_target_month3.xlsx", rows)
    summary = _import(path, "month3.xlsx")
    assert summary["skipped"] == 0

    db = SessionLocal()
    promoted = db.query(Village).filter(Village.village_code == "V-KEEP-2").one()
    assert promoted.target_classification == "هدف"
    db.close()
